"""Tests that generated trackers match the reference template's structure and styling."""

from pathlib import Path

import openpyxl
import pytest

from app.builder import TrackerData, TrackerItem, TrackerSection, build_tracker
from app.parser import parse_trace_csv

FIXTURES = Path(__file__).parent / "fixtures"
REFERENCE = FIXTURES / "IP_Mansfield_RB2_Tracksheet_2026.xlsx"


@pytest.fixture(scope="module")
def generated_workbook(tmp_path_factory):
    floor = parse_trace_csv(FIXTURES / "FLOOR.csv")
    front_wall_ports = parse_trace_csv(FIXTURES / "FRONT_WALL_W_PORTS.csv")
    front_wall_mlo = parse_trace_csv(FIXTURES / "FRONT_WALL_MLO.csv")

    data = TrackerData(
        title="IP MANSFIELD RB2 — 2026 OUTAGE NDE TRACKSHEET",
        customer=floor.company_name,
        location=floor.mill_location,
        equipment=floor.boiler_name,
        date=floor.inspection_date.strip(),
        sections=[
            TrackerSection(name=floor.boiler_section, elevations=[e.label for e in floor.elevations]),
            TrackerSection(
                name=front_wall_ports.boiler_section,
                elevations=[e.label for e in front_wall_ports.elevations],
            ),
            TrackerSection(
                name=front_wall_mlo.boiler_section,
                elevations=[e.label for e in front_wall_mlo.elevations],
            ),
        ],
    )

    out_dir = tmp_path_factory.mktemp("output")
    out_path = build_tracker(data, out_dir / "tracker.xlsx")
    return openpyxl.load_workbook(out_path)


def _font_tuple(font):
    color = font.color
    color_tuple = (color.type, color.rgb, color.theme, color.tint) if color else None
    return (font.name, font.sz, font.b, font.i, color_tuple)


def _fill_tuple(fill):
    if fill.fill_type is None:
        return None
    fg = fill.fgColor
    return (fill.fill_type, fg.type, fg.rgb, fg.theme, fg.tint)


def _border_tuple(border):
    return tuple((getattr(border, side).style if getattr(border, side) else None) for side in ("left", "right", "top", "bottom"))


def _alignment_tuple(alignment):
    return (alignment.horizontal, alignment.vertical, alignment.wrap_text)


def test_sheet_name_and_freeze_panes(generated_workbook):
    ws = generated_workbook["Tracker"]
    assert ws.freeze_panes == "A9"


def test_column_widths_match_reference(generated_workbook):
    ref = openpyxl.load_workbook(REFERENCE)["Tracker"]
    gen = generated_workbook["Tracker"]
    for letter in ("A", "J"):
        assert gen.column_dimensions[letter].width == pytest.approx(ref.column_dimensions[letter].width)


def test_columns_b_through_i_are_150px(generated_workbook):
    gen = generated_workbook["Tracker"]
    for letter in ("B", "C", "D", "E", "F", "G", "H", "I"):
        assert gen.column_dimensions[letter].width == pytest.approx(20.71)


def test_merged_cells(generated_workbook):
    ws = generated_workbook["Tracker"]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:A5" in merged
    assert "A7:J7" in merged
    for row in range(1, 5):
        assert f"E{row}:J{row}" in merged


def test_header_block_values(generated_workbook):
    ws = generated_workbook["Tracker"]
    assert ws["D1"].value == "Customer:"
    assert ws["E1"].value == "International Paper"
    assert ws["D2"].value == "Location:"
    assert ws["E2"].value == "Mansfield"
    assert ws["D3"].value == "Equipment:"
    assert ws["E3"].value == "Recovery Boiler #2"
    assert ws["D4"].value == "Project Date:"
    assert ws["E4"].value == "June 2026"
    assert ws["E4"].number_format == "@"
    assert ws["D5"].value is None
    assert ws["E5"].value == "Started"
    assert ws["D6"].value is None
    assert ws["E6"].value == "Complete"
    assert ws["A7"].value == "IP MANSFIELD RB2 — 2026 OUTAGE NDE TRACKSHEET"


def test_legend_cell_styling(generated_workbook):
    ws = generated_workbook["Tracker"]

    started = ws["E5"]
    assert started.font.bold is True
    assert started.font.color.rgb == "FF000000"
    assert started.fill.fill_type == "solid"
    assert started.fill.fgColor.rgb == "00FFFF00"

    complete = ws["E6"]
    assert complete.font.bold is True
    assert complete.font.color.rgb == "FF000000"
    assert complete.fill.fill_type == "solid"
    assert complete.fill.fgColor.rgb == "0000B050"


def test_column_headers_row8(generated_workbook):
    ws = generated_workbook["Tracker"]
    expected = [
        "Elevation", "Received", "Verifications Run", "Verifications Received",
        "Final Printed", "Wear Printed", "Forecasting Printed", "Trending Printed",
        "Exec Updated", "Notes",
    ]
    for col, heading in enumerate(expected, start=1):
        assert ws.cell(row=8, column=col).value == heading


@pytest.mark.parametrize("row", [7, 8, 9, 10, 19, 20, 28])
def test_styles_match_reference_for_shared_rows(generated_workbook, row):
    """Rows 1-40 of the generated file mirror the reference's FLOOR / FRONT WALL
    W/PORTS / FRONT WALL MLO sections, which are also the reference's first
    three sections in the same order."""
    ref = openpyxl.load_workbook(REFERENCE)["Tracker"]
    gen = generated_workbook["Tracker"]

    for col in range(1, 11):
        ref_cell = ref.cell(row=row, column=col)
        gen_cell = gen.cell(row=row, column=col)
        assert _font_tuple(gen_cell.font) == _font_tuple(ref_cell.font), f"font mismatch at row {row} col {col}"
        assert _fill_tuple(gen_cell.fill) == _fill_tuple(ref_cell.fill), f"fill mismatch at row {row} col {col}"
        if row == 8 and 6 <= col <= 9:
            # Reference omits the redundant right border on F8-I8 (the left
            # border of the next cell renders the same line); both are
            # visually identical so only compare left/top/bottom here.
            gen_b, ref_b = gen_cell.border, ref_cell.border
            assert (gen_b.left.style, gen_b.top.style, gen_b.bottom.style) == (
                ref_b.left.style, ref_b.top.style, ref_b.bottom.style
            ), f"border mismatch at row {row} col {col}"
            continue
        assert _border_tuple(gen_cell.border) == _border_tuple(ref_cell.border), f"border mismatch at row {row} col {col}"


def test_section_header_values(generated_workbook):
    ws = generated_workbook["Tracker"]
    assert ws["A9"].value == "FLOOR"
    assert ws["A20"].value == "FRONT WALL W/PORTS"
    assert ws["A28"].value == "FRONT WALL MLO"


def test_elevation_values(generated_workbook):
    ws = generated_workbook["Tracker"]
    assert ws["A10"].value == "-1'' FROM WELD"
    assert ws["A19"].value == "3X3 AREA AT SMELT SPOUTS 1'"


def test_closing_border_on_last_row(generated_workbook):
    ws = generated_workbook["Tracker"]
    last_row = ws.max_row
    for col in range(1, 11):
        cell = ws.cell(row=last_row, column=col)
        assert cell.border.bottom.style == "medium"


def _minimal_data(**overrides):
    base = dict(
        title="Test Tracker",
        customer="Test Co",
        location="Plant A",
        equipment="Boiler 1",
        date="June 2026",
        sections=[TrackerSection(name="FLOOR", elevations=["EL 1", "EL 2"])],
    )
    base.update(overrides)
    return TrackerData(**base)


def test_builder_omits_aux_section_when_empty(tmp_path):
    data = _minimal_data()
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    values = [ws.cell(row=r, column=1).value for r in range(9, ws.max_row + 1)]
    assert "AUXILIARY SCOPE ITEMS" not in values
    assert "PUNCHLIST" not in values


def test_builder_writes_aux_section_header(tmp_path):
    data = _minimal_data(
        auxiliary_items=[TrackerItem(description="PT OF COMPOSITE PORTS", notes="")]
    )
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    values = [ws.cell(row=r, column=1).value for r in range(9, ws.max_row + 1)]
    assert "AUXILIARY SCOPE ITEMS" in values


def test_builder_writes_punchlist_section_header(tmp_path):
    data = _minimal_data(
        punchlist_items=[TrackerItem(description='ITEM 37\nUT spout', notes='Reading: .286"')]
    )
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    values = [ws.cell(row=r, column=1).value for r in range(9, ws.max_row + 1)]
    assert "PUNCHLIST" in values


def test_builder_aux_item_description_and_notes(tmp_path):
    data = _minimal_data(
        auxiliary_items=[TrackerItem(description="RT OF ECONOMIZER", notes="Complete")]
    )
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    aux_row = next(
        r for r in range(9, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "AUXILIARY SCOPE ITEMS"
    )
    item_row = aux_row + 1
    assert ws.cell(row=item_row, column=1).value == "RT OF ECONOMIZER"
    assert ws.cell(row=item_row, column=10).value == "Complete"


def test_builder_closing_border_after_aux_and_punchlist(tmp_path):
    data = _minimal_data(
        auxiliary_items=[TrackerItem(description="AUX ITEM", notes="")],
        punchlist_items=[TrackerItem(description="PUNCH ITEM", notes="")],
    )
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    last_row = ws.max_row
    for col in range(1, 11):
        assert ws.cell(row=last_row, column=col).border.bottom.style == "medium"


def test_builder_aux_before_punchlist(tmp_path):
    data = _minimal_data(
        auxiliary_items=[TrackerItem(description="AUX ITEM", notes="")],
        punchlist_items=[TrackerItem(description="PUNCH ITEM", notes="")],
    )
    out = build_tracker(data, tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Tracker"]
    aux_row = next(r for r in range(9, ws.max_row + 1) if ws.cell(row=r, column=1).value == "AUXILIARY SCOPE ITEMS")
    punch_row = next(r for r in range(9, ws.max_row + 1) if ws.cell(row=r, column=1).value == "PUNCHLIST")
    assert aux_row < punch_row
