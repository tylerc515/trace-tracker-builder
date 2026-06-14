"""Post-generation validation: sanity-check generated tracker output files."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.builder import TrackerData
from app.constants import TRACKER_SHEET_NAME

HEADER_ROW_COUNT = 8


def validate_tracker_output(data: TrackerData, xlsx_path: Path, pdf_path: Optional[Path]) -> list[str]:
    """Re-check the generated files and return a list of warning messages, if any."""
    warnings: list[str] = []

    if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
        warnings.append(f"'{xlsx_path.name}' was not found or is empty after generation.")
        return warnings

    try:
        workbook = load_workbook(xlsx_path, read_only=True)
    except Exception:
        warnings.append(f"'{xlsx_path.name}' could not be reopened to verify its contents.")
        return warnings

    if TRACKER_SHEET_NAME not in workbook.sheetnames:
        warnings.append(f"The '{TRACKER_SHEET_NAME}' sheet is missing from '{xlsx_path.name}'.")
    else:
        worksheet = workbook[TRACKER_SHEET_NAME]
        expected_rows = HEADER_ROW_COUNT + sum(1 + len(section.elevations) for section in data.sections)
        if worksheet.max_row < expected_rows:
            warnings.append(
                f"'{xlsx_path.name}' has fewer rows than expected "
                f"({worksheet.max_row} of {expected_rows})."
            )
    workbook.close()

    if pdf_path is not None and (not pdf_path.exists() or pdf_path.stat().st_size == 0):
        warnings.append(f"'{pdf_path.name}' was not found or is empty after generation.")

    return warnings
