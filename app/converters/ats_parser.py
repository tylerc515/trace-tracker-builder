"""ATS xlsx inspection file parser."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


class ATSParseError(Exception):
    """Raised when an ATS xlsx file cannot be parsed."""


@dataclass
class ATSElevation:
    label: str
    tech_code: str
    nominal_wall: float | None
    left: list[str]
    cntr: list[str]
    rght: list[str]


@dataclass
class ATSParseResult:
    company_name: str
    mill_location: str
    boiler_name: str
    inspection_date: str       # "Month YYYY"
    boiler_section: str
    num_tubes: int
    numbering_direction: str   # "Left-to-Right" or "Right-to-Left"
    nde_laboratory: str
    year: int
    ats_flags: dict[str, str]  # {code: description}
    elevations: list[ATSElevation]
    tube_numbers: list[int]


_MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def convert_reading(value: object) -> str:
    """Convert an ATS cell value to Standard Format string.

    float/int -> zero-padded 3-digit integer (0.234 -> "234", 0.010 -> "010")
    None or empty string -> ""
    str (flag code) -> unchanged
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{round(value * 1000):03d}"
    stripped = str(value).strip()
    return stripped


def _parse_flag_legend(raw: object) -> dict[str, str]:
    if not isinstance(raw, str) or not raw.strip():
        return {}
    result: dict[str, str] = {}
    for part in raw.split(";"):
        part = part.strip()
        if " - " in part:
            code, desc = part.split(" - ", 1)
            result[code.strip()] = desc.strip()
    return result


def _parse_direction(raw: object) -> str:
    if isinstance(raw, str):
        upper = raw.upper()
        if "LEFT TO RIGHT" in upper:
            return "Left-to-Right"
        if "RIGHT TO LEFT" in upper:
            return "Right-to-Left"
    return "Left-to-Right"


def _parse_date(raw: object) -> str:
    if not isinstance(raw, str):
        return ""
    match = re.search(r"Date:\s*(\d{1,2})/(\d{2})/(\d{4})", raw)
    if not match:
        return ""
    month_num = int(match.group(1))
    year = match.group(3)
    if 1 <= month_num <= 12:
        return f"{_MONTH_NAMES[month_num]} {year}"
    return ""


def parse_ats_file(filepath: str | Path) -> ATSParseResult:
    """Parse an ATS Thickness xlsx file into an ATSParseResult.

    Raises:
        ATSParseError: if the file cannot be opened or is missing expected data.
    """
    path = Path(filepath)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise ATSParseError(f"Cannot open '{path.name}': {exc}") from exc

    if "Thickness" not in wb.sheetnames:
        raise ATSParseError(f"'{path.name}' has no 'Thickness' sheet")

    ws = wb["Thickness"]

    # Row 1 (openpyxl row 1): year and flag legend
    year_cell = ws.cell(row=1, column=1).value
    year = int(year_cell) if isinstance(year_cell, (int, float)) else 0
    ats_flags = _parse_flag_legend(ws.cell(row=1, column=3).value)

    # Row 2 (openpyxl row 2): numbering direction
    numbering_direction = _parse_direction(ws.cell(row=2, column=3).value)

    # Row 3 (openpyxl row 3): tube numbers starting at col D (column 4)
    tube_numbers: list[int] = []
    col = 4
    while True:
        val = ws.cell(row=3, column=col).value
        if val is None:
            break
        if isinstance(val, str) and val.upper().strip() in ("AVG", "MIN"):
            break
        if isinstance(val, (int, float)):
            tube_numbers.append(int(val))
            col += 1
        else:
            break

    num_tubes = len(tube_numbers)
    if num_tubes == 0:
        raise ATSParseError(f"'{path.name}': no tube numbers found in row 3")

    # Elevation blocks starting at openpyxl row 4
    elevations: list[ATSElevation] = []
    row_idx = 4
    while row_idx <= ws.max_row:
        col_a = ws.cell(row=row_idx, column=1).value
        col_c = ws.cell(row=row_idx, column=3).value

        # Stop at repeat footer header
        if col_a == "ELEVATION" and col_c == "LOC":
            break

        if col_c == "L":
            tech_raw = ws.cell(row=row_idx, column=1).value
            tech_code = str(tech_raw).strip() if tech_raw is not None else ""

            nominal_raw = ws.cell(row=row_idx + 1, column=1).value
            nominal_wall = (
                float(nominal_raw)
                if isinstance(nominal_raw, (int, float))
                else None
            )

            label_parts = []
            for sub_row in (row_idx, row_idx + 1, row_idx + 2):
                part = ws.cell(row=sub_row, column=2).value
                if part is not None:
                    stripped = str(part).strip()
                    if stripped:
                        label_parts.append(stripped)
            label = " ".join(label_parts)

            left = [
                convert_reading(ws.cell(row=row_idx, column=4 + i).value)
                for i in range(num_tubes)
            ]
            cntr = [
                convert_reading(ws.cell(row=row_idx + 1, column=4 + i).value)
                for i in range(num_tubes)
            ]
            rght = [
                convert_reading(ws.cell(row=row_idx + 2, column=4 + i).value)
                for i in range(num_tubes)
            ]

            elevations.append(ATSElevation(
                label=label,
                tech_code=tech_code,
                nominal_wall=nominal_wall,
                left=left,
                cntr=cntr,
                rght=rght,
            ))
            row_idx += 3
        else:
            row_idx += 1

    # Metadata from bottom-right area
    company_name = mill_location = boiler_name = boiler_section = ""
    nde_laboratory = ""
    inspection_date = ""

    scan_start = max(1, ws.max_row - 40)
    anchor_row: int | None = None
    anchor_col: int | None = None

    for ri in range(scan_start, ws.max_row + 1):
        for ci in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=ri, column=ci).value
            if isinstance(cell_val, str) and "TUBE WALL THICKNESS SURVEY" in cell_val:
                anchor_row = ri
                anchor_col = ci
                break
        if anchor_row is not None:
            break

    if anchor_row is not None and anchor_col is not None:
        def _str_at(r: int) -> str:
            v = ws.cell(row=r, column=anchor_col).value
            return str(v).strip() if v is not None else ""

        company_name = _str_at(anchor_row - 2)
        mill_location = _str_at(anchor_row - 1)
        boiler_name = _str_at(anchor_row + 1)
        boiler_section = _str_at(anchor_row + 2)

    for ri in range(scan_start, ws.max_row + 1):
        for ci in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=ri, column=ci).value
            if isinstance(cell_val, str) and "\n" in cell_val and not nde_laboratory:
                nde_laboratory = cell_val.split("\n")[0].strip().title()
            if isinstance(cell_val, str) and "Date:" in cell_val and not inspection_date:
                inspection_date = _parse_date(cell_val)

    if not company_name:
        raise ATSParseError(f"'{path.name}': could not locate metadata block")

    return ATSParseResult(
        company_name=company_name,
        mill_location=mill_location,
        boiler_name=boiler_name,
        inspection_date=inspection_date,
        boiler_section=boiler_section,
        num_tubes=num_tubes,
        numbering_direction=numbering_direction,
        nde_laboratory=nde_laboratory,
        year=year,
        ats_flags=ats_flags,
        elevations=elevations,
        tube_numbers=tube_numbers,
    )
