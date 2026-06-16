"""Excel tracker generation matching the reference TRACE tracker template.

Styles below (fonts, fills, borders, column widths, freeze panes) were
extracted by inspecting IP_Mansfield_RB2_Tracksheet_2026.xlsx with openpyxl
and are reproduced here so generated workbooks match it exactly.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import TRACKER_COLUMNS, TRACKER_SHEET_NAME

# --- Column layout -----------------------------------------------------

TRACKER_COLUMN_COUNT = len(TRACKER_COLUMNS)  # A through J
LAST_COLUMN_LETTER = "J"

COLUMN_WIDTHS = {
    "A": 44.1796875,
    "B": 20.71,
    "C": 20.71,
    "D": 20.71,
    "E": 20.71,
    "F": 20.71,
    "G": 20.71,
    "H": 20.71,
    "I": 20.71,
    "J": 65.1796875,
}

HEADER_ROW_HEIGHTS = {1: 15.75, 2: 15.75, 3: 15.75, 4: 15.75, 7: 19.5, 8: 31.5}
SECTION_ROW_HEIGHT = 21.0
ELEVATION_ROW_HEIGHT = 15.0

# --- Fonts ---------------------------------------------------------------

_TEXT_COLOR = Color(theme=1, tint=0.0)

LABEL_FONT = Font(name="Calibri", size=12, bold=True, color=_TEXT_COLOR)
VALUE_FONT = Font(name="Calibri", size=12, bold=False, color=_TEXT_COLOR)
LEGEND_FONT = Font(name="Calibri", size=11, bold=False, color=_TEXT_COLOR)
LEGEND_FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=Color(rgb="FF000000"))
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=_TEXT_COLOR)
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color=_TEXT_COLOR)
SECTION_FONT = Font(name="Calibri", size=16, bold=True, color=_TEXT_COLOR)
ELEVATION_FONT = Font(name="Calibri", size=11, bold=False, color=_TEXT_COLOR)

# --- Alignment -------------------------------------------------------------

LABEL_ALIGNMENT = Alignment(vertical="center")
VALUE_ALIGNMENT = Alignment(horizontal="left", vertical="center")
TITLE_ALIGNMENT = Alignment(horizontal="left", vertical="center")
HEADER_ALIGNMENT_NOWRAP = Alignment(horizontal="center", vertical="center")
HEADER_ALIGNMENT_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ELEVATION_LABEL_ALIGNMENT = Alignment(horizontal="center")
PLACEHOLDER_ALIGNMENT = Alignment(horizontal="center")

# --- Fill ------------------------------------------------------------------

SECTION_FILL = PatternFill(fill_type="solid", fgColor=Color(theme=0, tint=-0.249977111117893))
LEGEND_STARTED_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
LEGEND_COMPLETE_FILL = PatternFill(fill_type="solid", fgColor="00B050")

# --- BSI logo ----------------------------------------------------------

BSI_LOGO_FILENAME = "bsi_logo.jpg"
BSI_LOGO_MAX_WIDTH = 130
BSI_LOGO_MAX_HEIGHT = 60

# --- Borders -----------------------------------------------------------

_THIN = Side(style="thin", color=Color(indexed=64))
_MEDIUM = Side(style="medium", color=Color(indexed=64))
_SIDES = {"thin": _THIN, "medium": _MEDIUM, None: None}


def _border(left=None, right=None, top=None, bottom=None) -> Border:
    return Border(left=_SIDES[left], right=_SIDES[right], top=_SIDES[top], bottom=_SIDES[bottom])


def _with_bottom(border: Border, style: str) -> Border:
    return Border(left=border.left, right=border.right, top=border.top, bottom=_SIDES[style])


# --- Data model --------------------------------------------------------

@dataclass
class TrackerSection:
    """A single boiler section and the elevations it contains, in order."""

    name: str
    elevations: list[str] = field(default_factory=list)


@dataclass
class TrackerItem:
    """A single auxiliary scope or punchlist entry."""

    description: str
    notes: str = ""


@dataclass
class TrackerData:
    """Everything needed to render the Tracker sheet."""

    title: str
    customer: str
    location: str
    equipment: str
    date: str
    sections: list[TrackerSection] = field(default_factory=list)
    auxiliary_items: list[TrackerItem] = field(default_factory=list)
    punchlist_items: list[TrackerItem] = field(default_factory=list)


def _write_header_block(ws: Worksheet, data: TrackerData) -> None:
    """Write rows 1-8: title block, summary fields, and column headers."""
    ws.merge_cells("A1:A5")
    ws["A1"].alignment = Alignment(horizontal="center")

    summary_rows = [
        (1, "Customer:", data.customer, _border(bottom="thin")),
        (2, "Location:", data.location, _border(top="thin", bottom="thin")),
        (3, "Equipment:", data.equipment, _border(top="thin", bottom="thin")),
        (4, "Project Date:", data.date, _border(top="thin")),
    ]
    for row, label, value, value_border in summary_rows:
        label_cell = ws.cell(row=row, column=4, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LABEL_ALIGNMENT

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
        value_cell = ws.cell(row=row, column=5, value=value)
        value_cell.font = VALUE_FONT
        value_cell.alignment = VALUE_ALIGNMENT
        if row == 4:
            value_cell.number_format = "@"
        for col in range(5, 11):
            ws.cell(row=row, column=col).border = value_border

    # Started / Complete legend
    for col in range(5, 11):
        ws.cell(row=5, column=col).border = _border(bottom="thin")
    legend_started = ws.cell(row=5, column=5, value="Started")
    legend_started.font = LEGEND_FONT_BOLD
    legend_started.fill = LEGEND_STARTED_FILL
    legend_started.border = _border(left="thin", right="thin", top="thin", bottom="thin")

    for col in range(5, 11):
        ws.cell(row=6, column=col).border = _border(top="thin")
    legend_complete = ws.cell(row=6, column=5, value="Complete")
    legend_complete.font = LEGEND_FONT_BOLD
    legend_complete.fill = LEGEND_COMPLETE_FILL
    legend_complete.border = _border(left="thin", right="thin", top="thin", bottom="thin")

    # Title row
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=10)
    title_cell = ws.cell(row=7, column=1, value=data.title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGNMENT
    for col in range(1, 11):
        left = "medium" if col == 1 else None
        right = "medium" if col == TRACKER_COLUMN_COUNT else None
        ws.cell(row=7, column=col).border = _border(left=left, right=right, top="thin", bottom="medium")

    # Column header row
    for col, heading in enumerate(TRACKER_COLUMNS, start=1):
        cell = ws.cell(row=8, column=col, value=heading)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT_NOWRAP if col <= 2 else HEADER_ALIGNMENT_WRAP
        left = "medium" if col == 1 else "thin"
        right = "medium" if col == TRACKER_COLUMN_COUNT else "thin"
        cell.border = _border(left=left, right=right, top="medium", bottom="medium")

    for row, height in HEADER_ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height


def _write_sections(ws: Worksheet, sections: list[TrackerSection]) -> int:
    """Write each section header and its elevation rows. Returns the last row used."""
    row = 9
    for section in sections:
        for col in range(1, TRACKER_COLUMN_COUNT + 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.value = section.name
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            left = "medium" if col == 1 else "thin"
            right = "medium" if col == TRACKER_COLUMN_COUNT else "thin"
            cell.border = _border(left=left, right=right, top="medium", bottom="medium")
            if col == TRACKER_COLUMN_COUNT:
                cell.alignment = HEADER_ALIGNMENT_WRAP
        ws.row_dimensions[row].height = SECTION_ROW_HEIGHT
        row += 1

        elevation_count = len(section.elevations)
        for index, elevation in enumerate(section.elevations):
            if elevation_count == 1:
                top, bottom = None, None
            elif index == 0:
                top, bottom = None, "thin"
            elif index == elevation_count - 1:
                top, bottom = "thin", None
            else:
                top, bottom = "thin", "thin"

            for col in range(1, TRACKER_COLUMN_COUNT + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = ELEVATION_FONT
                left = "medium" if col == 1 else "thin"
                right = "medium" if col == TRACKER_COLUMN_COUNT else "thin"
                cell.border = _border(left=left, right=right, top=top, bottom=bottom)
                if col == 1:
                    cell.value = elevation
                    cell.alignment = ELEVATION_LABEL_ALIGNMENT
            ws.row_dimensions[row].height = ELEVATION_ROW_HEIGHT
            row += 1

    return row - 1


def _write_extra_sections(ws: Worksheet, label: str, items: list[TrackerItem], start_row: int) -> int:
    """Write an AUXILIARY SCOPE ITEMS or PUNCHLIST block. Returns the last row written."""
    row = start_row

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TRACKER_COLUMN_COUNT)
    for col in range(1, TRACKER_COLUMN_COUNT + 1):
        cell = ws.cell(row=row, column=col)
        if col == 1:
            cell.value = label
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        left = "medium" if col == 1 else "thin"
        right = "medium" if col == TRACKER_COLUMN_COUNT else "thin"
        cell.border = _border(left=left, right=right, top="medium", bottom="medium")
    ws.row_dimensions[row].height = SECTION_ROW_HEIGHT
    row += 1

    for item in items:
        for col in range(1, TRACKER_COLUMN_COUNT + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ELEVATION_FONT
            left = "medium" if col == 1 else "thin"
            right = "medium" if col == TRACKER_COLUMN_COUNT else "thin"
            cell.border = _border(left=left, right=right, top="thin", bottom="thin")
            if col == 1:
                cell.value = item.description
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col == TRACKER_COLUMN_COUNT:
                cell.value = item.notes if item.notes else None
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    return row - 1


def _apply_column_widths(ws: Worksheet) -> None:
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _bsi_logo_path() -> Path:
    """Return the path to bsi_logo.jpg, whether running from source or frozen."""
    if getattr(sys, "frozen", False):
        base_dir = Path(sys._MEIPASS)  # type: ignore[attr-defined]
    else:
        base_dir = Path(__file__).resolve().parent.parent
    return base_dir / BSI_LOGO_FILENAME


def _add_bsi_logo(ws: Worksheet) -> None:
    """Embed the BSI logo in the top-left of the worksheet, anchored to A1."""
    logo_path = _bsi_logo_path()
    if not logo_path.exists():
        return

    image = XLImage(str(logo_path))
    scale = min(BSI_LOGO_MAX_WIDTH / image.width, BSI_LOGO_MAX_HEIGHT / image.height)
    image.width = image.width * scale
    image.height = image.height * scale
    ws.add_image(image, "A1")


def build_tracker(data: TrackerData, output_path: str | Path) -> Path:
    """Generate the Tracker workbook for the given data and save it to output_path."""
    wb = Workbook()
    ws = wb.active
    ws.title = TRACKER_SHEET_NAME

    _write_header_block(ws, data)
    _add_bsi_logo(ws)
    last_row = _write_sections(ws, data.sections)

    if data.auxiliary_items:
        last_row = _write_extra_sections(ws, "AUXILIARY SCOPE ITEMS", data.auxiliary_items, last_row + 1)
    if data.punchlist_items:
        last_row = _write_extra_sections(ws, "PUNCHLIST", data.punchlist_items, last_row + 1)

    # Close the table with a medium bottom border on the final row.
    if last_row >= 9:
        for col in range(1, TRACKER_COLUMN_COUNT + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.border = _with_bottom(cell.border, "medium")

    _apply_column_widths(ws)
    ws.freeze_panes = "A9"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
